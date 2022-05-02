from tkinter import *
from tkinter.ttk import *
from openpyxl import workbook, load_workbook
import random
root = Tk()
root.title("Randomization lottery")
root.geometry("300x300")
root.resizable(0,0)


wb = load_workbook("hello.xlsx")
ws = wb.active
rangeline = ws["A2": "A19"]
name = []
for items in rangeline:
    for subitems in items:
        name.append(subitems.value)

computer_action = random.choice(name)
print("The computer randomly chose: " + computer_action)

mainloop()