import random
from openpyxl import workbook
from openpyxl import load_workbook
import tkinter as tk
from tkinter.ttk import *
import pandas as pd
from PIL import Image, ImageTk
# creating tkinter window
root = tk.Tk()
root.title('CSUMB Application')
root.geometry("600x600")
root.resizable(0,0)
root.grid_columnconfigure(1, weight=1)
photo=tk.PhotoImage(file="logofinal.png")
myimage=Label(image=photo)
myimage.grid(row=0, column=1, padx=0, pady= 20)


#first button that gets data from excel file CSUMB
def apple():
   data = pd.read_excel("exam1.xlsx")
   df = pd.DataFrame(data, columns=['CalendarDate'])
   lb.config(text=df)
b1 = tk.Button(text="Calendar", font=("Arial", 15), bg="#A3E4D7", fg= "blue",
              command=apple)
b1.place(x=60, y=150)
lb = tk.Label( font=("Arial", 10), text=" ", justify="left")
lb.place(x=125, y=180)

def apple1():
   data1 = pd.read_excel("exam1.xlsx")
   col = pd.DataFrame(data1, columns=['Buildings'])
   lb.config(text=col)
b2 = tk.Button(text="Buildings", font=("Arial", 15), bg="#A3E4D7",fg= "blue",
              command= apple1)
b2.place(x=170, y=150)
lb = tk.Label( font=("Arial", 10), text=" ", justify="left")
lb.place(x=30, y=180)

def apple2():
   data2 = pd.read_excel("exam1.xlsx")
   fac = pd.DataFrame(data2, columns=['FacultyName'])
   lb.config(text=fac)
b3 = tk.Button(text="Faculty", font=("Arial", 15), bg="#A3E4D7",fg= "blue",
              command= apple2)
b3.place(x=270, y=150)
lb = tk.Label( font=("Arial", 10), text=" ", justify="left")
lb.place(x=30, y=180)

def randnum():
    wb = load_workbook('exam1.xlsx')
    ws = wb.active
    range = ws['C2':'C13']
    l = []
    for cell in range:
        for x in cell:
            l.append(x.value)
    computer_action = random.choice(l)
    print(computer_action)
    l1.config(text=computer_action)
b4 = tk.Button(text="Select Faculty Randomly", font=("Arial", 15), bg="#A3E4D7",fg= "blue",
               command=randnum)
b4.place(x=360, y=150)
l1 = tk.Label(font=("Arial", 10))
l1.place(x=30, y=180)
#b1.place(x=110, y=70)
#l1.place(x=165, y=130)
tk.mainloop()
